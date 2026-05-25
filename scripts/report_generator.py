"""
report_generator.py
1. Lee el JSON de detecciones del día y el del día anterior.
2. Llama a Claude para generar un resumen inteligente en español.
3. Construye un email HTML bonito.
4. Lo envía a santicomet@gmail.com vía Gmail SMTP.
5. Guarda un Excel acumulativo con todas las detecciones.
"""

import json
import os
import smtplib
import ssl
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR  = Path("data")
REPORT_DIR = Path("reports")
REPORT_DIR.mkdir(exist_ok=True)

ACCOUNT_NAMES = [n.strip() for n in os.environ.get("ACCOUNT_NAMES", "").split(",")]
REPORT_EMAIL  = os.environ.get("REPORT_EMAIL", "santicomet@gmail.com")
GMAIL_SENDER  = os.environ.get("GMAIL_SENDER", "")
GMAIL_APP_PWD = os.environ.get("GMAIL_APP_PASSWORD", "")


# ── Carga de datos ───────────────────────────────────────────────────────────
def load_json(date_str: str) -> list:
    path = DATA_DIR / f"detections_{date_str}.json"
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return []


def compute_diff(today: list, yesterday: list) -> dict:
    """Compara detecciones de hoy vs ayer por cuenta."""
    yday_counts = {r["account"]: len(r["detections"]) for r in yesterday}
    diff = {}
    for r in today:
        acc   = r["account"]
        cnt   = len(r["detections"])
        prev  = yday_counts.get(acc, 0)
        diff[acc] = {"today": cnt, "yesterday": prev, "delta": cnt - prev, "error": r.get("error")}
    return diff


# ── Resumen con Claude ───────────────────────────────────────────────────────
def aggregate_all_detections(today: list) -> list:
    """Une todas las detecciones de las 8 cuentas en una sola lista (mismo artista)."""
    all_detections = []
    for r in today:
        for d in r["detections"]:
            d["_account_slot"] = r["account"]  # guardamos el slot para debug
            all_detections.append(d)
    return all_detections


def get_top_channels(detections: list, n=5) -> list:
    from collections import Counter
    counts = Counter(d.get("channel", "Desconocido") for d in detections)
    return counts.most_common(n)


def get_top_tracks(detections: list, n=5) -> list:
    from collections import Counter
    counts = Counter(d.get("title", "Desconocido") for d in detections)
    return counts.most_common(n)


def generate_ai_summary(today: list, diff: dict) -> str:
    client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))

    all_detections = aggregate_all_detections(today)
    total_today    = len(all_detections)
    total_yesterday = sum(d["yesterday"] for d in diff.values())
    delta = total_today - total_yesterday

    top_channels = get_top_channels(all_detections, 5)
    top_tracks   = get_top_tracks(all_detections, 5)
    errors       = [r["account"] for r in today if r.get("error")]

    context = {
        "fecha": datetime.now().strftime("%d/%m/%Y"),
        "total_detecciones_hoy": total_today,
        "total_detecciones_ayer": total_yesterday,
        "cambio_vs_ayer": delta,
        "canales_mas_activos": [{"canal": c, "detecciones": n} for c, n in top_channels],
        "canciones_mas_detectadas": [{"cancion": t, "veces": n} for t, n in top_tracks],
        "paises": list({d.get("country","") for d in all_detections if d.get("country")}),
        "cuentas_con_error": errors,
        "muestra_detecciones": all_detections[:5],
    }

    prompt = f"""Eres el asistente personal de Santi Comet, músico y compositor español.
Todas las detecciones son de su catálogo propio — las 8 cuentas de Tunesat son simplemente 
particiones técnicas (la versión gratuita solo permite 50 canciones por cuenta).

Datos consolidados de hoy:
{json.dumps(context, ensure_ascii=False, indent=2)}

Escribe un resumen diario en español, tono profesional pero cercano (tutéale).
El resumen debe:
1. Abrir con el total de detecciones y si sube o baja respecto a ayer
2. Mencionar los canales de TV/radio más activos hoy
3. Si hay canciones concretas destacadas, mencionarlas
4. Mencionar países si hay diversidad geográfica interesante
5. Si hay errores en algún slot, mencionarlo brevemente al final
6. Cerrar con una frase motivadora muy breve

Máximo 160 palabras. Sin asteriscos, sin markdown, solo texto limpio en párrafos."""

    message = client.messages.create(
        model="claude-sonnet-4-5-20251001",
        max_tokens=400,
        messages=[{"role": "user", "content": prompt}]
    )
    return message.content[0].text


# ── Excel acumulativo ────────────────────────────────────────────────────────
def update_excel(today: list, date_str: str) -> Path:
    excel_path = DATA_DIR / "historico_detecciones.xlsx"

    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Detecciones"
        headers = ["Fecha", "Cuenta", "Canal", "Canción", "Duración", "País", "Hora"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF", name="Arial")
            cell.fill = PatternFill("solid", start_color="1A1A2E")
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        widths = [12, 14, 20, 30, 10, 12, 10]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # Añadir filas nuevas
    thin = Side(style="thin", color="CCCCDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    next_row = ws.max_row + 1

    for r in today:
        for d in r["detections"]:
            row_data = [
                date_str,
                r["account"],
                d.get("channel", ""),
                d.get("title", ""),
                d.get("duration", ""),
                d.get("country", ""),
                d.get("time", ""),
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=next_row, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal="center" if col != 4 else "left")
                cell.font = Font(name="Arial", size=10)
            next_row += 1

    wb.save(excel_path)
    return excel_path


# ── HTML del email ───────────────────────────────────────────────────────────
def build_html_email(today: list, diff: dict, ai_summary: str, date_str: str) -> str:
    from collections import Counter

    date_fmt = datetime.strptime(date_str, "%Y-%m-%d").strftime("%d de %B de %Y")

    all_detections  = aggregate_all_detections(today)
    total_today     = len(all_detections)
    total_yesterday = sum(d["yesterday"] for d in diff.values())
    delta_total     = total_today - total_yesterday
    delta_str       = f"+{delta_total}" if delta_total >= 0 else str(delta_total)
    delta_color     = "#27ae60" if delta_total >= 0 else "#e74c3c"

    # Top canales
    top_channels = Counter(d.get("channel","Desconocido") for d in all_detections).most_common(5)
    channels_html = ""
    max_cnt = top_channels[0][1] if top_channels else 1
    for ch, cnt in top_channels:
        pct = int(cnt / max_cnt * 100)
        channels_html += f"""
        <tr>
          <td style="padding:7px 12px;color:#333;font-size:13px">{ch}</td>
          <td style="padding:7px 12px;width:160px">
            <div style="background:#e8e8f5;border-radius:4px;height:8px">
              <div style="background:#1A1A2E;width:{pct}%;height:8px;border-radius:4px"></div>
            </div>
          </td>
          <td style="padding:7px 12px;text-align:right;font-weight:700;color:#1A1A2E;font-size:13px">{cnt}</td>
        </tr>"""

    # Top canciones
    top_tracks = Counter(d.get("title","Desconocido") for d in all_detections).most_common(5)
    tracks_html = ""
    for i, (track, cnt) in enumerate(top_tracks, 1):
        tracks_html += f"""
        <tr style="border-bottom:1px solid #eee">
          <td style="padding:8px 12px;color:#888;font-size:12px">#{i}</td>
          <td style="padding:8px 12px;color:#1A1A2E;font-weight:600;font-size:13px">{track}</td>
          <td style="padding:8px 12px;text-align:right;color:#555;font-size:13px">{cnt}x</td>
        </tr>"""

    # Distribución geográfica
    countries = Counter(d.get("country","") for d in all_detections if d.get("country"))
    geo_html = " &nbsp;·&nbsp; ".join(
        f'<span style="font-weight:600;color:#1A1A2E">{c}</span> <span style="color:#888">({n})</span>'
        for c, n in countries.most_common(8)
    ) or '<span style="color:#888">Sin datos de país</span>'

    # Tabla detalle completa (todas las detecciones unificadas, ordenadas por hora)
    sorted_detections = sorted(all_detections, key=lambda d: d.get("time",""), reverse=True)
    detail_rows = ""
    for d in sorted_detections[:50]:
        detail_rows += f"""
        <tr style="border-bottom:1px solid #f0f0f8">
          <td style="padding:6px 10px;color:#555;font-size:12px">{d.get('time','')}</td>
          <td style="padding:6px 10px;color:#1A1A2E;font-weight:600;font-size:12px">{d.get('title','')}</td>
          <td style="padding:6px 10px;color:#333;font-size:12px">{d.get('channel','')}</td>
          <td style="padding:6px 10px;color:#666;font-size:12px;text-align:center">{d.get('country','')}</td>
          <td style="padding:6px 10px;color:#888;font-size:12px;text-align:center">{d.get('duration','')}</td>
        </tr>"""
    if len(all_detections) > 50:
        detail_rows += f'<tr><td colspan="5" style="padding:10px;text-align:center;color:#888;font-style:italic;font-size:12px">... y {len(all_detections)-50} detecciones más en el Excel adjunto</td></tr>'

    # Errores de slots
    errors_html = ""
    slot_errors = [(r["account"], r["error"]) for r in today if r.get("error")]
    if slot_errors:
        errors_html = '<div style="background:#fff5f5;border:1px solid #fcc;border-radius:8px;padding:12px 16px;margin:0 32px 24px">'
        errors_html += '<p style="margin:0 0 6px;font-weight:600;color:#c0392b;font-size:13px">⚠️ Slots con error</p>'
        for slot, err in slot_errors:
            errors_html += f'<p style="margin:2px 0;font-size:12px;color:#666"><b>{slot}:</b> {err}</p>'
        errors_html += '</div>'

    html = f"""<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#f2f3f8;font-family:Arial,sans-serif">
<div style="max-width:660px;margin:30px auto;background:#fff;border-radius:14px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,.1)">

  <!-- Header -->
  <div style="background:linear-gradient(135deg,#1A1A2E 0%,#2d2d5e 100%);padding:32px;text-align:center">
    <div style="font-size:32px;margin-bottom:8px">🎵</div>
    <h1 style="color:#fff;margin:0;font-size:24px;letter-spacing:2px;font-weight:700">SANTI COMET</h1>
    <p style="color:#8899cc;margin:6px 0 0;font-size:13px;letter-spacing:1px">TUNESAT DAILY REPORT · {date_fmt.upper()}</p>
  </div>

  <!-- Resumen IA -->
  <div style="background:#f8f9ff;padding:22px 32px;border-bottom:2px solid #e8e8f5">
    <p style="margin:0;color:#333;line-height:1.75;font-size:14px">{ai_summary}</p>
  </div>

  <!-- KPIs principales -->
  <div style="padding:24px 32px 8px;display:flex;gap:12px">
    <div style="flex:1;background:#1A1A2E;border-radius:10px;padding:18px;text-align:center">
      <div style="font-size:36px;font-weight:700;color:#fff">{total_today}</div>
      <div style="font-size:11px;color:#8899cc;margin-top:4px;letter-spacing:1px">DETECCIONES HOY</div>
    </div>
    <div style="flex:1;background:#f0f0f8;border-radius:10px;padding:18px;text-align:center">
      <div style="font-size:36px;font-weight:700;color:{delta_color}">{delta_str}</div>
      <div style="font-size:11px;color:#888;margin-top:4px;letter-spacing:1px">VS AYER ({total_yesterday})</div>
    </div>
    <div style="flex:1;background:#f0f0f8;border-radius:10px;padding:18px;text-align:center">
      <div style="font-size:36px;font-weight:700;color:#1A1A2E">{len(countries)}</div>
      <div style="font-size:11px;color:#888;margin-top:4px;letter-spacing:1px">PAÍSES</div>
    </div>
  </div>

  <!-- Distribución geográfica -->
  <div style="padding:16px 32px 8px">
    <p style="margin:0;font-size:12px;color:#888">📍 {geo_html}</p>
  </div>

  <!-- Top canales -->
  <div style="padding:20px 32px 8px">
    <h2 style="font-size:14px;color:#1A1A2E;margin:0 0 12px;text-transform:uppercase;letter-spacing:1px">📺 Canales más activos</h2>
    <table width="100%" cellpadding="0" cellspacing="0">
      {channels_html if channels_html else '<tr><td style="color:#888;font-size:13px;padding:8px 0">Sin datos de canal hoy.</td></tr>'}
    </table>
  </div>

  <!-- Top canciones -->
  <div style="padding:20px 32px 8px">
    <h2 style="font-size:14px;color:#1A1A2E;margin:0 0 12px;text-transform:uppercase;letter-spacing:1px">🎼 Canciones más emitidas</h2>
    <table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse">
      {tracks_html if tracks_html else '<tr><td style="color:#888;font-size:13px;padding:8px 0">Sin datos de canciones hoy.</td></tr>'}
    </table>
  </div>

  <!-- Tabla detalle -->
  <div style="padding:20px 32px 24px">
    <h2 style="font-size:14px;color:#1A1A2E;margin:0 0 12px;text-transform:uppercase;letter-spacing:1px">📋 Todas las detecciones</h2>
    {'<table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse"><tr style="background:#1A1A2E"><th style="padding:8px 10px;text-align:left;color:#8899cc;font-size:11px;font-weight:500">HORA</th><th style="padding:8px 10px;text-align:left;color:#8899cc;font-size:11px;font-weight:500">CANCIÓN</th><th style="padding:8px 10px;text-align:left;color:#8899cc;font-size:11px;font-weight:500">CANAL</th><th style="padding:8px 10px;text-align:center;color:#8899cc;font-size:11px;font-weight:500">PAÍS</th><th style="padding:8px 10px;text-align:center;color:#8899cc;font-size:11px;font-weight:500">DUR.</th></tr>' + detail_rows + '</table>' if all_detections else '<p style="color:#888;font-size:13px">Sin detecciones nuevas hoy.</p>'}
  </div>

  {errors_html}

  <!-- Footer -->
  <div style="background:#1A1A2E;padding:16px 32px;text-align:center">
    <p style="color:#4455778;margin:0;font-size:11px;color:#556699">
      Generado automáticamente por Tunesat Agent · Santi Comet © {datetime.now().year}<br>
      400 canciones monitorizadas · 8 slots · Tunesat.com
    </p>
  </div>
</div>
</body>
</html>"""
    return html


# ── Envío de email ───────────────────────────────────────────────────────────
def send_email(html: str, excel_path: Path, date_str: str):
    msg = MIMEMultipart("mixed")
    msg["Subject"] = f"🎵 Tunesat Daily Report · {date_str}"
    msg["From"]    = GMAIL_SENDER
    msg["To"]      = REPORT_EMAIL

    # El HTML va dentro de un contenedor "alternative" anidado (buena práctica MIME)
    body = MIMEMultipart("alternative")
    body.attach(MIMEText(html, "html", "utf-8"))
    msg.attach(body)

    # Adjuntar Excel
    if excel_path.exists():
        with open(excel_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="Tunesat_Historico_{date_str}.xlsx"')
        msg.attach(part)

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(GMAIL_SENDER, GMAIL_APP_PWD)
        server.sendmail(GMAIL_SENDER, REPORT_EMAIL, msg.as_string())

    print(f"✉️  Email enviado a {REPORT_EMAIL}")


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    date_str   = datetime.now().strftime("%Y-%m-%d")
    yesterday  = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")

    print("📊 Generando reporte consolidado de Santi Comet...")
    today_data = load_json(date_str)
    prev_data  = load_json(yesterday)

    if not today_data:
        print("⚠️  No hay datos para hoy. ¿Ejecutó el scraper correctamente?")
        return

    all_det = aggregate_all_detections(today_data)
    print(f"   Total detecciones hoy: {len(all_det)} (de {len(today_data)} slots)")

    diff       = compute_diff(today_data, prev_data)
    ai_summary = generate_ai_summary(today_data, diff)
    excel_path = update_excel(today_data, date_str)
    html       = build_html_email(today_data, diff, ai_summary, date_str)

    # Guardar HTML del reporte
    report_path = REPORT_DIR / f"report_{date_str}.html"
    report_path.write_text(html, encoding="utf-8")

    send_email(html, excel_path, date_str)
    print("✅ Reporte completado.")


if __name__ == "__main__":
    main()
